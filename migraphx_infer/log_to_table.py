import os
import pandas as pd
import argparse
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment

parser = argparse.ArgumentParser()
parser.add_argument("--directory", required=True)
args = parser.parse_args()

data = []
meta_col = ["Model name", "Error"]
data_col = [
    "Session create time",
    "First inference time",
    "Total inference time",
    "Average inference time",
    "Inferences/second",
    "Avg CPU usage",
    "Peak working set size",
    "Min Latency",
    "Max Latency",
    "P50 Latency",
    "P90 Latency",
    "P95 Latency",
    "P99 Latency",
    "P999 Latency",
]
data_str = [
    "Session creation time cost:",
    "First inference time cost:",
    "Total inference time cost:",
    "Average inference time cost:",
    "Number of inferences per second:",
    "Avg CPU usage:",
    "Peak working set size:",
    "Min Latency:",
    "Max Latency:",
    "P50 Latency:",
    "P90 Latency:",
    "P95 Latency:",
    "P99 Latency:",
    "P999 Latency:",
]

error_keyword = ["timed out after", "died with", "MIGraphX Error", "Run failed", "Fatal error", "Type Error"]
def contains_stat(lines):
    for line in lines:
        if data_str[1] in line:
            return True


def parse_perf(lines):
    ret = [""] * len(data_str)
    for line in lines:
        for i in range(len(data_str)):
            data_type = data_str[i]
            if data_type in line:
                res = line.split(data_type)[1].strip()
                ret[i] = res
    return ret

def parse_error(lines):
    for line in lines:
        for i in range(len(error_keyword)):
            err_type = error_keyword[i]
            if err_type in line:
                res = line.split(err_type)[1]
                return err_type + res
    return "Unclassified error"


for filename in os.listdir(args.directory):
    if filename.endswith(".log"):
        file_path = os.path.join(args.directory, filename)
        with open(file_path, "r", encoding="utf-8") as f:
            txt = f.read().strip()
            model_name = filename.replace(".log", ".onnx")
            lines = txt.split("\n")
            if contains_stat(lines):
                stat = parse_perf(lines)
                data.append([model_name, "N/A", *stat])
            else:
                error_msg = parse_error(lines)
                data.append([model_name, error_msg])

df = pd.DataFrame(data, columns=[*meta_col, *data_col])

output_excel = "output.xlsx"
df.to_excel(output_excel, index=False)

wb = load_workbook(output_excel)
ws = wb.active

# Find the "Error" column letter
error_col_letter = None
for cell in ws[1]:
    if str(cell.value).strip().lower() == "error":
        error_col_letter = cell.column_letter
        break

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass

    if column == error_col_letter:
        adjusted_width = min(max_length + 2, 50)
    else:
        adjusted_width = max_length + 2

    ws.column_dimensions[column].width = adjusted_width
    

wb.save(output_excel)
